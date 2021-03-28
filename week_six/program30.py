import pandas
from openpyxl import Workbook
from openpyxl.styles import PatternFill

all_employees = pandas.read_csv("zamci_vykazy.csv", index_col="cislo_zamestnance")
print(all_employees.head(), "\n")

# Ulož tabulku s cekovými počty vykázaných hodin, kterou jsi vytvořila v příkladu 27 jako Excel soubor.
# Výsledný sešit si otevři a zkontroluj. K uložení použij funkci to_excel,
# se kterou pracuj stejně, jako jsme na lekci pracovali s funkci to_csv.
all_employees.to_excel("zamci_vykazy.xlsx")

# Zkus data z Excelu naopak načíst pomocí funkci read_excel() a ověř, že se soubor načetl v pořádku.
print(pandas.read_excel("zamci_vykazy.xlsx", index_col="cislo_zamestnance"))

# Zkus pomocí modulu zapsat rozvrh hodin jako tabulku v Excel sešitu.
# Níže máš program, který obsahuje rozvrh hodin jako dvourozměrný seznam.
# Vnitřní seznamy obsahují předměty v rozvrhu, jeden vnitřní seznam vždy obsahuje předměty pro daný den.
wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"

schedule = [
  ["Anglický jazyk", "Přírodopis", "Dějepis", "Matematika", "Oběd", "Tělesná výchova", "Tělesná výchova", ],
  ["Občanská výchova", "Informatika", "Matematika", "Oběd", "Výtvarná výchova", "Dějepis", ],
  ["Matematika", "Chemie", "Přírodopis", "Fyzika", "Oběd", "Zeměpis", ],
  ["Fyzika", "Anglický jazyk", "Matematika", "Český jazyk", "Dějepis", "Oběd", ],
  ["Český jazyk", "Zeměpis", "Český jazyk", "Datová Analýza", "Oběd", ]
]

blue = PatternFill("solid", fgColor="BCE0E2")
green = PatternFill("solid", fgColor="E2BEBC")
purple = PatternFill("solid", fgColor="D0A3EB")
yellow = PatternFill("solid", fgColor="FFFECC")

row = 1
for day in schedule:
  cell_class = ws1.cell(row, 1)
  cell_class.fill = blue
  col = 1
  for school_class in day:
    cell_class = ws1.cell(row, col)
    cell_class.value = school_class
    if cell_class.value == "Oběd":
      cell_class.fill = green
    elif cell_class.value == "Informatika":
      cell_class.fill = yellow
    elif cell_class.value == "Datová Analýza":
      cell_class.fill = purple
    col += 1
  row += 1

wb.save(filename="rozvrh_hodin.xlsx")

# Podívej se nejprve na ukázku.
# Jednoduchý program níže zapíše hodnotu Test do buňky B1 a výsledek uloží souboru rozvrh_hodin.xlsx.
# Pokud neznáš terminologii Excelu, pak Workbook označuje sešit,
# tj. celý "soubor". ws se odkazuje na Work Sheet, což je list, tj. jedna "záložka".
# Náš soubor bude mít jen jeden list, čímž je situace jednoduchá.

